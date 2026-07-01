"""
db_loader.py — parses Navy Recruiting Sheet (High School Players tab).
CRITICAL: Always use parse_xlsx() with the xlsx file. Never rely on Drive text.

parse_xlsx(path) reads all 1000+ rows directly from Excel with full fidelity.
parse_sheet_content(text) is a fallback for Drive markdown only.

Column map (0-indexed from xlsx, LOCKED — verified against the real header
row 2026-06-30, do not guess, re-verify against row 3 of the live sheet if
this ever looks wrong):
  [7]=First [8]=Last [9]=Class [10]=★/tier [11]=Commit [12]=Pos
  [16]=State [17]=High School [18]=Summer Team [22]=Seen
Row numbers (1-indexed, matching openpyxl/Sheets) are tracked per entry as
'_row' — needed by sheet_write.py to target updates at the right row.
"""

NICKNAMES = {
    'chris':['christopher'], 'christopher':['chris'],
    'jake':['jacob'],        'jacob':['jake'],
    'will':['william'],      'william':['will'],
    'mike':['michael'],      'michael':['mike'],
    'matt':['matthew'],      'matthew':['matt'],
    'alex':['alexander'],    'alexander':['alex'],
    'zach':['zachary'],      'zachary':['zach'],
    'joe':['joseph'],        'joseph':['joe'],
    'nick':['nicholas'],     'nicholas':['nick'],
    'ben':['benjamin'],      'benjamin':['ben'],
    'sam':['samuel'],        'samuel':['sam'],
    'dan':['daniel'],        'daniel':['dan'],
    'rob':['robert'],        'robert':['rob'],
    'pat':['patrick'],       'patrick':['pat'],
    'drew':['andrew'],       'andrew':['drew'],
    'cal':['caleb'],         'caleb':['cal'],
    'conor':['connor'],      'connor':['conor'],
    'tom':['thomas'],        'thomas':['tom'],
    'tim':['timothy'],       'timothy':['tim'],
    'jim':['james'],         'james':['jim'],
    'brad':['brady'],        'brady':['brad'],
    # Additional common baseball name variants
    'jon':['jonathan','johnathan'], 'jonathan':['jon'], 'johnathan':['jon'],
    'nate':['nathaniel','nathan'],  'nathaniel':['nate'], 'nathan':['nate'],
    'cj':['christopher','charles'], 'ty':['tyler','tyson'],
    'eli':['elijah'],               'elijah':['eli'],
    'jj':['james','john'],
    'trey':['tre'],                 'tre':['trey'],
    'cole':['coleman','nicolas'],
    'max':['maxwell','maximilian'], 'maxwell':['max'],
    'bj':['brian','brandon'],
    'aj':['andrew','anthony'],
    'tj':['thomas','tyler'],
    'ry':['ryan'],
    'zeke':['ezekiel'],             'ezekiel':['zeke'],
    'luke':['lucas'],               'lucas':['luke'],
    'cam':['cameron'],              'cameron':['cam'],
    'cade':['caden','caiden'],      'caden':['cade'], 'caiden':['cade'],
    'jay':['jason','james','jayden'],
    'ryan':['ry'],
    'liam':['william'],
    'jack':['jackson','john'],      'jackson':['jack'],
}

# Suffixes to strip before any name matching
_SUFFIXES = {'jr', 'jr.', 'sr', 'sr.', 'ii', 'iii', 'iv', 'v',
             'the third', 'the second', '2nd', '3rd', '4th'}

def strip_suffix(name):
    """Remove trailing generational suffixes: Jr., III, etc."""
    parts = name.lower().strip().split()
    while parts and parts[-1].rstrip('.') in _SUFFIXES:
        parts.pop()
    return ' '.join(parts)

def _variants(first, last):
    first = strip_suffix(first.lower().strip())
    last  = strip_suffix(last.lower().strip())
    v = {f"{first} {last}"}
    for alt in NICKNAMES.get(first, []):
        v.add(f"{alt} {last}")
    return v

# Header text -> canonical field name. Resolved from the LIVE header row at
# parse time, never hardcoded column numbers - those broke silently on
# 2026-06-30 when Chris deleted two leading columns and every hardcoded
# index quietly pointed at the wrong field. If the sheet's headers are ever
# renamed, update the labels below - do not go back to fixed indices.
_HEADER_LABELS = {
    'id': 'ID', 'name': 'Name', 'pos_group': 'Pos Group',
    'date_added': 'Date Added:', 'by': 'By:',
    'first': 'First Name', 'last': 'Last Name', 'class': 'Class',
    'tier': '★', 'commit': 'Commit', 'pos': 'Pos',
    'state': 'State', 'hs': 'High School', 'team': 'Summer Team',
    'seen': 'Seen', 'notes': 'Notes',
}


def find_columns(path, sheet_name='High School Players', header_row=3):
    """Resolve current 1-indexed column numbers by scanning the live header
    row. Raises clearly if a needed header is missing rather than silently
    reading/writing the wrong column."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header = {}
    for cell in ws[header_row]:
        if cell.value not in (None, ''):
            header[str(cell.value).strip()] = cell.column
    wb.close()
    cols, missing = {}, []
    for field, label in _HEADER_LABELS.items():
        if label in header:
            cols[field] = header[label]
        else:
            missing.append(label)
    if missing:
        raise ValueError(f"Recruiting sheet header changed - could not find column(s) "
                         f"{missing} in row {header_row}. Check the sheet's real header "
                         f"before trusting any read or write.")
    return cols


def parse_xlsx(path, sheet_name='High School Players'):
    import openpyxl
    cols = find_columns(path, sheet_name)
    i_first, i_last, i_class = cols['first']-1, cols['last']-1, cols['class']-1
    i_tier, i_commit, i_pos = cols['tier']-1, cols['commit']-1, cols['pos']-1
    i_state, i_hs, i_team, i_seen = cols['state']-1, cols['hs']-1, cols['team']-1, cols['seen']-1
    max_idx = max(i_first, i_last, i_class, i_tier, i_commit, i_pos, i_state, i_hs, i_team, i_seen)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    db = {}
    skipped = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 3: continue
        if len(row) <= max_idx: continue
        first  = str(row[i_first]).strip()  if row[i_first]  is not None else ''
        last   = str(row[i_last]).strip()   if row[i_last]   is not None else ''
        yr     = str(row[i_class]).strip()  if row[i_class]  is not None else ''
        tier   = str(row[i_tier]).strip()   if row[i_tier]   is not None else ''
        commit = str(row[i_commit]).strip() if row[i_commit] is not None else ''
        pos    = str(row[i_pos]).strip()    if row[i_pos]    is not None else ''
        state  = str(row[i_state]).strip()  if row[i_state]  is not None else ''
        hs     = str(row[i_hs]).strip()     if row[i_hs]     is not None else ''
        team   = str(row[i_team]).strip()   if row[i_team]   is not None else ''
        seen   = str(row[i_seen]).strip()   if row[i_seen]   is not None else ''
        if yr.endswith('.0'): yr = yr[:-2]
        if tier.endswith('.0') and tier != '0.1': tier = tier[:-2]
        if not first or not last: skipped += 1; continue
        if first in ('First Name','First',':-:','By:','NAME'): skipped += 1; continue
        if not yr or not yr.startswith('20'): skipped += 1; continue
        entry = {'tier':tier,'pos':pos,'class':yr,'commit':commit,
                 'first':first,'last':last,'canonical_name':f"{first} {last}",
                 'state':state,'hs':hs,'team':team,'seen':seen,'_row':i}
        for v in _variants(first, last):
            db[v] = entry
    wb.close()
    unique = len(set(e['canonical_name'] for e in db.values()))
    print(f"[DB] Loaded {unique} players from xlsx ({len(db)} name variants)")
    return db

def parse_sheet_content(raw_text):
    """Fallback for Drive markdown text — less complete, use only if no xlsx."""
    db = {}
    for line in raw_text.split('\n'):
        if not line.startswith('|') or ':-:' in line: continue
        cols = [c.strip() for c in line.split('|')]
        if len(cols) < 12: continue
        first = cols[8] if len(cols)>8 else ''
        last  = cols[9] if len(cols)>9 else ''
        yr    = cols[10] if len(cols)>10 else ''
        tier  = cols[11] if len(cols)>11 else ''
        commit= cols[12] if len(cols)>12 else ''
        pos   = cols[13] if len(cols)>13 else ''
        if not first or not last: continue
        if first in ('First Name',':-:','By:','NAME','TEAM','First'): continue
        if not yr or not yr.startswith('20'): continue
        entry = {'tier':tier,'pos':pos,'class':yr,'commit':commit,
                 'first':first,'last':last,'canonical_name':f"{first} {last}"}
        for v in _variants(first, last):
            db[v] = entry
    unique = len(set(e['canonical_name'] for e in db.values()))
    print(f"[DB] Loaded {unique} players from sheet text ({len(db)} name variants)")
    return db

def lookup(db, scraped_name):
    if not scraped_name: return None
    # Strip suffix first (Jr., III, etc.)
    cleaned = strip_suffix(scraped_name)
    key = cleaned.strip()
    if key in db: return db[key]
    parts = key.split()
    if len(parts) >= 2:
        first, last = parts[0], ' '.join(parts[1:])
        for alt in NICKNAMES.get(first, []):
            if f"{alt} {last}" in db: return db[f"{alt} {last}"]
    # Fuzzy fallback — last name must be exact, fuzzy only on first name
    return _fuzzy_lookup(db, key)

def _fuzzy_lookup(db, cleaned_name, threshold=0.82):
    """Last-resort: exact last name + fuzzy first name match. High threshold only."""
    from difflib import SequenceMatcher
    parts = cleaned_name.split()
    if len(parts) < 2: return None
    first_in = parts[0]
    last_in  = ' '.join(parts[1:])
    best, best_score = None, 0
    for key, entry in db.items():
        k_parts = key.split()
        if len(k_parts) < 2: continue
        k_first = k_parts[0]
        k_last  = ' '.join(k_parts[1:])
        if k_last != last_in: continue   # last name must be exact
        score = SequenceMatcher(None, first_in, k_first).ratio()
        if score > best_score:
            best, best_score = entry, score
    if best_score >= threshold:
        print(f"[DB] Fuzzy match: '{cleaned_name}' → '{best['canonical_name']}' ({best_score:.2f})")
        return best
    return None
